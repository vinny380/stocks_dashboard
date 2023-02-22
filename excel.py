import openpyxl
import streamlit as st
from io import BytesIO
import time

output = BytesIO()

excel_file = st.file_uploader('Upload your excel file')
@st.cache_data()
def modify(excel_file):
    wb = openpyxl.load_workbook(BytesIO(excel_file.read()))
    sheet = wb.active #specify the sheet name to select other than the active sheet
    print("Maximum rows before removing : ",sheet.max_row)
    sheet.insert_cols(1, 2)
    sheet.insert_cols(idx=4)
    sheet.delete_rows(0,17)
    sheet['A1'] = "Subject Number"
    sheet['B1'] = 'list'
    sheet['D1'] = 'null'
    for row in sheet.iter_rows():
        values = [cell.value for cell in row]
        values[4], values[10] = values[10], values[4]
        for i, cell in enumerate(row):
            cell.value = values[i]

    print("Maximum rows after removing : ", sheet.max_row)
    path1 = 'edited_.xlsx'
    wb.save(path1)
    time.sleep(1)
    st.download_button(
        label="Download Updated Excel Workbook",
        data=open(path1, 'rb').read(),
        file_name="workbook.xlsx",
        mime="xlsx"
    )
# 10 to 

edit = st.button('Click me')
if edit:
    modify(excel_file)